from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.http import HttpResponse
from django.db import transaction
from .models import CuentaContable
from .serializers import CuentaContableSerializer, CuentaContableTreeSerializer
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class CuentaContableViewSet(viewsets.ModelViewSet):
    """ViewSet básico para gestión de cuentas contables MVP"""
    serializer_class = CuentaContableSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo', 'naturaleza', 'nivel', 'afectable']
    search_fields = ['codigo', 'nombre']
    ordering_fields = ['codigo', 'nombre', 'nivel']
    ordering = ['codigo']
    
    def get_queryset(self):
        """Solo cuentas de la empresa actual"""
        # Intentar obtener empresa desde middleware
        if hasattr(self.request, 'empresa') and self.request.empresa:
            return CuentaContable.objects.filter(
                empresa=self.request.empresa,
                activo=True
            ).select_related('cuenta_padre')
        
        # Si no hay middleware, obtener desde sesión o primera empresa del usuario
        empresa_id = self.request.session.get('empresa_id')
        if empresa_id:
            return CuentaContable.objects.filter(
                empresa_id=empresa_id,
                activo=True
            ).select_related('cuenta_padre')
        
        # Como fallback, usar la primera empresa del usuario
        from apps.empresas.models import UsuarioEmpresa
        acceso = UsuarioEmpresa.objects.filter(
            usuario=self.request.user,
            activo=True
        ).first()
        if acceso:
            return CuentaContable.objects.filter(
                empresa=acceso.empresa,
                activo=True
            ).select_related('cuenta_padre')
            
        return CuentaContable.objects.none()
        
    def perform_create(self, serializer):
        """Asignar empresa actual al crear"""
        if hasattr(self.request, 'empresa') and self.request.empresa:
            serializer.save(
                empresa=self.request.empresa,
                creado_por=self.request.user
            )
        else:
            # Si no hay empresa seleccionada, usar la primera del usuario
            from apps.empresas.models import UsuarioEmpresa
            acceso = UsuarioEmpresa.objects.filter(
                usuario=self.request.user,
                activo=True
            ).first()
            if acceso:
                serializer.save(
                    empresa=acceso.empresa,
                    creado_por=self.request.user
                )
            else:
                raise Exception("No tiene acceso a ninguna empresa")
    
    def perform_update(self, serializer):
        """Actualizar modificado_por"""
        serializer.save(modificado_por=self.request.user)
    
    @action(detail=False, methods=['get'])
    def arbol(self, request):
        """Vista en árbol de cuentas contables"""
        # Usar la misma lógica que get_queryset para obtener la empresa
        empresa = None
        if hasattr(request, 'empresa') and request.empresa:
            empresa = request.empresa
        else:
            empresa_id = request.session.get('empresa_id')
            if empresa_id:
                from apps.empresas.models import Empresa
                try:
                    empresa = Empresa.objects.get(id=empresa_id)
                except Empresa.DoesNotExist:
                    pass
            
            if not empresa:
                from apps.empresas.models import UsuarioEmpresa
                acceso = UsuarioEmpresa.objects.filter(
                    usuario=request.user,
                    activo=True
                ).first()
                if acceso:
                    empresa = acceso.empresa
        
        if empresa:
            # Solo cuentas de primer nivel
            cuentas_mayor = CuentaContable.objects.filter(
                empresa=empresa,
                cuenta_padre__isnull=True,
                activo=True
            ).order_by('codigo')
            
            serializer = CuentaContableTreeSerializer(cuentas_mayor, many=True)
            return Response(serializer.data)
        return Response([])
    
    @action(detail=True, methods=['get'])
    def subcuentas(self, request, pk=None):
        """Obtiene subcuentas de una cuenta específica"""
        cuenta = self.get_object()
        subcuentas = cuenta.subcuentas.filter(activo=True).order_by('codigo')
        serializer = CuentaContableSerializer(subcuentas, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def tipos(self, request):
        """Lista los tipos de cuenta disponibles"""
        tipos = CuentaContable._meta.get_field('tipo').choices
        return Response([
            {'valor': valor, 'nombre': nombre} 
            for valor, nombre in tipos
        ])
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exportar catálogo de cuentas a Excel"""
        # Obtener empresa
        empresa = None
        if hasattr(request, 'empresa') and request.empresa:
            empresa = request.empresa
        else:
            empresa_id = request.session.get('empresa_id')
            if empresa_id:
                from apps.empresas.models import Empresa
                try:
                    empresa = Empresa.objects.get(id=empresa_id)
                except Empresa.DoesNotExist:
                    pass
            
            if not empresa:
                from apps.empresas.models import UsuarioEmpresa
                acceso = UsuarioEmpresa.objects.filter(
                    usuario=request.user,
                    activo=True
                ).first()
                if acceso:
                    empresa = acceso.empresa
        
        if not empresa:
            return Response({'error': 'Sin empresa seleccionada'}, status=400)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo de Cuentas"
        
        # Headers con estilo
        headers = ['Código', 'Nombre', 'Tipo', 'Naturaleza', 'Nivel', 'Cuenta Padre', 'Afectable', 'Descripción']
        
        # Aplicar estilos a headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Obtener cuentas ordenadas
        cuentas = CuentaContable.objects.filter(
            empresa=empresa,
            activo=True
        ).order_by('codigo').select_related('cuenta_padre')
        
        # Llenar datos
        for row, cuenta in enumerate(cuentas, 2):
            ws.cell(row=row, column=1, value=cuenta.codigo)
            ws.cell(row=row, column=2, value=cuenta.nombre)
            ws.cell(row=row, column=3, value=cuenta.tipo)
            ws.cell(row=row, column=4, value=cuenta.naturaleza)
            ws.cell(row=row, column=5, value=cuenta.nivel)
            ws.cell(row=row, column=6, value=cuenta.cuenta_padre.codigo if cuenta.cuenta_padre else '')
            ws.cell(row=row, column=7, value='Sí' if cuenta.afectable else 'No')
            ws.cell(row=row, column=8, value=cuenta.descripcion or '')
        
        # Ajustar ancho de columnas
        column_widths = [15, 40, 15, 15, 8, 15, 12, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="catalogo_cuentas_{empresa.codigo}.xlsx"'
        
        # Guardar workbook
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def template_excel(self, request):
        """Descargar template de Excel para importación"""
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Cuentas"
        
        # Headers
        headers = ['Código', 'Nombre', 'Tipo', 'Naturaleza', 'Código Cuenta Padre', 'Afectable', 'Descripción']
        
        # Aplicar estilos a headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Agregar ejemplos
        ejemplos = [
            ['1', 'ACTIVO', 'ACTIVO', 'DEUDORA', '', 'No', 'Cuenta mayor de activos'],
            ['1.1', 'ACTIVO CIRCULANTE', 'ACTIVO', 'DEUDORA', '1', 'No', 'Activos de corto plazo'],
            ['1.1.1', 'CAJA', 'ACTIVO', 'DEUDORA', '1.1', 'Sí', 'Efectivo en caja'],
        ]
        
        for row, ejemplo in enumerate(ejemplos, 2):
            for col, valor in enumerate(ejemplo, 1):
                ws.cell(row=row, column=col, value=valor)
        
        # Crear hoja de instrucciones
        ws_inst = wb.create_sheet("Instrucciones")
        instrucciones = [
            "INSTRUCCIONES PARA IMPORTAR CATÁLOGO DE CUENTAS",
            "",
            "1. Llenar la hoja 'Template Cuentas' con sus datos",
            "2. Campos requeridos: Código, Nombre, Tipo, Naturaleza",
            "3. Tipos válidos: ACTIVO, PASIVO, CAPITAL, INGRESO, GASTO", 
            "4. Naturaleza válida: DEUDORA, ACREEDORA",
            "5. Código Cuenta Padre debe existir previamente",
            "6. Afectable: Sí/No (indica si permite movimientos)",
            "",
            "IMPORTANTE:",
            "- Los códigos deben ser únicos",
            "- Crear cuentas padre antes que subcuentas",
            "- Guardar como .xlsx antes de importar"
        ]
        
        for row, instruccion in enumerate(instrucciones, 1):
            cell = ws_inst.cell(row=row, column=1, value=instruccion)
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif instruccion.startswith(("IMPORTANTE", "1.", "2.", "3.", "4.", "5.", "6.")):
                cell.font = Font(bold=True)
        
        # Ajustar ancho de columnas
        column_widths = [15, 40, 15, 15, 20, 12, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        ws_inst.column_dimensions['A'].width = 60
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_catalogo_cuentas.xlsx"'
        
        # Guardar workbook
        wb.save(response)
        return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_cuenta_ajax(request):
    """Crear cuenta contable vía AJAX desde el frontend"""
    if request.method == 'POST':
        try:
            # Obtener empresa actual
            empresa_id = request.session.get('empresa_id')
            if not empresa_id:
                from apps.empresas.models import UsuarioEmpresa
                acceso = UsuarioEmpresa.objects.filter(
                    usuario=request.user,
                    activo=True
                ).first()
                if acceso:
                    empresa = acceso.empresa
                else:
                    return Response({'error': 'Sin empresa seleccionada'}, status=400)
            else:
                from apps.empresas.models import Empresa
                empresa = Empresa.objects.get(id=empresa_id)
            
            # Obtener datos del request
            data = request.data
            
            # Obtener cuenta padre si se especifica
            cuenta_padre = None
            if data.get('cuenta_padre'):
                cuenta_padre = CuentaContable.objects.get(
                    id=data['cuenta_padre'],
                    empresa=empresa
                )
            
            # Calcular nivel automáticamente
            nivel = 1 if not cuenta_padre else cuenta_padre.nivel + 1
            
            # Crear cuenta
            nueva_cuenta = CuentaContable.objects.create(
                empresa=empresa,
                codigo=data['codigo'],
                nombre=data['nombre'],
                tipo=data['tipo'],
                naturaleza=data['naturaleza'],
                nivel=nivel,
                cuenta_padre=cuenta_padre,
                afectable=data.get('afectable', True),
                descripcion=data.get('descripcion', ''),
                creado_por=request.user
            )
            
            return Response({
                'id': nueva_cuenta.id,
                'codigo': nueva_cuenta.codigo,
                'nombre': nueva_cuenta.nombre,
                'tipo': nueva_cuenta.tipo,
                'nivel': nueva_cuenta.nivel,
                'mensaje': f'Cuenta {nueva_cuenta.codigo} creada exitosamente'
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=400)
    
    return Response({'error': 'Método no permitido'}, status=405)


@api_view(['GET'])
@permission_classes([AllowAny])
def catalogo_root(request):
    """Página de navegación del catálogo de cuentas"""
    if 'text/html' in request.META.get('HTTP_ACCEPT', ''):
        html_content = """
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>Catálogo de Cuentas - ZivaBSuite</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; background: #f8f9fa; }
                .container { max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; }
                h1 { color: #2c3e50; }
                .endpoint { margin: 15px 0; padding: 15px; background: #f1f2f6; border-radius: 5px; }
                .method { font-weight: bold; color: #27ae60; }
                .url { color: #3498db; font-family: monospace; }
                .back-link { margin-bottom: 20px; }
                .back-link a { color: #7f8c8d; text-decoration: none; }
                .warning { background: #fff3cd; border: 1px solid #ffeaa7; padding: 10px; border-radius: 5px; margin: 15px 0; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="back-link"><a href="/">← Volver al inicio</a> | <a href="/api/">API Root</a></div>
                <h1>📊 Catálogo de Cuentas</h1>
                
                <div class="warning">
                    <strong>⚠️ Nota:</strong> Los endpoints marcados requieren autenticación (login admin o token API)
                </div>
                
                <h2>🔗 Endpoints Disponibles</h2>
                
                <div class="endpoint">
                    <div class="method">GET</div>
                    <div class="url"><a href="/api/catalogo/cuentas/">/api/catalogo/cuentas/</a></div>
                    <p>Lista todas las cuentas contables (🔒 Requiere autenticación)</p>
                </div>
                
                <div class="endpoint">
                    <div class="method">GET</div>
                    <div class="url"><a href="/api/catalogo/cuentas/arbol/">/api/catalogo/cuentas/arbol/</a></div>
                    <p>Vista en árbol del catálogo de cuentas (🔒 Requiere autenticación)</p>
                </div>
                
                <div class="endpoint">
                    <div class="method">GET</div>
                    <div class="url"><a href="/api/catalogo/cuentas/tipos/">/api/catalogo/cuentas/tipos/</a></div>
                    <p>Lista los tipos de cuenta disponibles (🔒 Requiere autenticación)</p>
                </div>
                
                <h2>🔍 Ejemplos de Filtros</h2>
                <ul>
                    <li><code>/api/catalogo/cuentas/?tipo=ACTIVO</code> - Solo cuentas de activo</li>
                    <li><code>/api/catalogo/cuentas/?nivel=1</code> - Solo cuentas mayor</li>
                    <li><code>/api/catalogo/cuentas/?search=CAJA</code> - Buscar por nombre</li>
                </ul>
                
                <h2>📋 Gestión Administrativa</h2>
                <p><a href="/admin/catalogo_cuentas/cuentacontable/">🛡️ Administración de Cuentas</a></p>
                
                <h2>🔑 Autenticación</h2>
                <p>Para usar los endpoints desde herramientas como curl o Postman:</p>
                <p><code>Authorization: Token 1e0b0f1c08f1359f4c76e55c2fcba894976aeba7</code></p>
            </div>
        </body>
        </html>
        """
        return HttpResponse(html_content)
    
    # Para requests JSON
    return Response({
        'message': 'Catálogo de Cuentas API',
        'version': '3.0.0-MVP',
        'endpoints': {
            'cuentas': request.build_absolute_uri('/api/catalogo/cuentas/'),
            'arbol': request.build_absolute_uri('/api/catalogo/cuentas/arbol/'),
            'tipos': request.build_absolute_uri('/api/catalogo/cuentas/tipos/'),
        },
        'auth_required': True,
        'token': 'Authorization: Token 1e0b0f1c08f1359f4c76e55c2fcba894976aeba7'
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def aplicar_plantilla_predefinida(request):
    """Aplicar catálogo predefinido según tipo de empresa"""
    tipo_plantilla = request.data.get('tipo')
    
    # Mapeo de plantillas
    PLANTILLAS = {
        'servicios': {
            'nombre': 'Empresa de Servicios',
            'cuentas': [
                ('1010', 'Caja', 'ACTIVO', 'DEUDORA', 'Efectivo y equivalentes'),
                ('1020', 'Bancos', 'ACTIVO', 'DEUDORA', 'Efectivo y equivalentes'),
                ('1030', 'Inversiones temporales', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('1040', 'Cuentas por cobrar clientes', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1050', 'Anticipos a proveedores', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1060', 'Gastos pagados por anticipado', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1070', 'Propiedad, planta y equipo', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('1080', 'Activos intangibles', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('2010', 'Proveedores', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('2020', 'Acreedores diversos', 'PASIVO', 'ACREEDORA', 'Operación/Financiación'),
                ('2030', 'Impuestos por pagar', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('2040', 'Préstamos bancarios corto plazo', 'PASIVO', 'ACREEDORA', 'Financiación'),
                ('2050', 'Cuentas por pagar empleados', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('2060', 'Provisiones', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('3010', 'Capital social', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3020', 'Reservas legales', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3030', 'Resultados acumulados', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3040', 'Resultado del ejercicio', 'CAPITAL', 'ACREEDORA', 'Operación'),
                ('4010', 'Ingresos por servicios', 'INGRESO', 'ACREEDORA', 'Operación'),
                ('4020', 'Ingresos por mantenimiento', 'INGRESO', 'ACREEDORA', 'Operación'),
                ('5010', 'Costo de servicios prestados', 'GASTO', 'DEUDORA', 'Operación'),
                ('5020', 'Gastos de personal', 'GASTO', 'DEUDORA', 'Operación'),
                ('5030', 'Gastos de oficina', 'GASTO', 'DEUDORA', 'Operación'),
                ('5040', 'Gastos de tecnología', 'GASTO', 'DEUDORA', 'Operación'),
                ('5050', 'Gastos de ventas', 'GASTO', 'DEUDORA', 'Operación'),
                ('5060', 'Gastos administrativos', 'GASTO', 'DEUDORA', 'Operación'),
                ('5070', 'Gastos financieros', 'GASTO', 'DEUDORA', 'Operación/Financiación'),
                ('5080', 'Depreciación y amortización', 'GASTO', 'DEUDORA', 'Operación (Ajuste no efectivo)'),
            ]
        },
        'comercial': {
            'nombre': 'Empresa Comercializadora',
            'cuentas': [
                ('1010', 'Caja y bancos', 'ACTIVO', 'DEUDORA', 'Efectivo y equivalentes'),
                ('1020', 'Inventarios de mercancías', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1030', 'Cuentas por cobrar clientes', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1040', 'Deudores diversos', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1050', 'Almacenes y equipos', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('1060', 'Vehículos de reparto', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('1070', 'Mobiliario y equipo oficina', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('2010', 'Proveedores', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('2020', 'Documentos por pagar', 'PASIVO', 'ACREEDORA', 'Financiación'),
                ('2030', 'Acreedores diversos', 'PASIVO', 'ACREEDORA', 'Operación/Financiación'),
                ('2040', 'Impuestos por pagar', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('2050', 'Préstamos bancarios', 'PASIVO', 'ACREEDORA', 'Financiación'),
                ('3010', 'Capital social', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3020', 'Utilidades retenidas', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3030', 'Resultado del ejercicio', 'CAPITAL', 'ACREEDORA', 'Operación'),
                ('4010', 'Ventas de mercancías', 'INGRESO', 'ACREEDORA', 'Operación'),
                ('4020', 'Devoluciones en ventas', 'INGRESO', 'DEUDORA', 'Operación'),
                ('4030', 'Descuentos comerciales', 'INGRESO', 'DEUDORA', 'Operación'),
                ('4040', 'Ingresos financieros', 'INGRESO', 'ACREEDORA', 'Operación'),
                ('5010', 'Costo de ventas', 'COSTO', 'DEUDORA', 'Operación'),
                ('5020', 'Gastos de venta', 'GASTO', 'DEUDORA', 'Operación'),
                ('5030', 'Gastos de administración', 'GASTO', 'DEUDORA', 'Operación'),
                ('5040', 'Gastos de distribución', 'GASTO', 'DEUDORA', 'Operación'),
                ('5050', 'Gastos financieros', 'GASTO', 'DEUDORA', 'Operación/Financiación'),
                ('5060', 'Depreciación', 'GASTO', 'DEUDORA', 'Operación (Ajuste no efectivo)'),
            ]
        },
        'industrial': {
            'nombre': 'Empresa de Manufactura',
            'cuentas': [
                ('1010', 'Efectivo y equivalentes', 'ACTIVO', 'DEUDORA', 'Efectivo y equivalentes'),
                ('1021', 'Inventario - Materias primas', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1022', 'Inventario - Productos en proceso', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1023', 'Inventario - Productos terminados', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1030', 'Cuentas por cobrar', 'ACTIVO', 'DEUDORA', 'Operación'),
                ('1041', 'Maquinaria', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('1042', 'Edificios', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('1043', 'Equipos de producción', 'ACTIVO', 'DEUDORA', 'Inversión'),
                ('2010', 'Proveedores materias primas', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('2020', 'Acreedores varios', 'PASIVO', 'ACREEDORA', 'Operación/Financiación'),
                ('2030', 'Impuestos por pagar', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('2040', 'Préstamos a corto plazo', 'PASIVO', 'ACREEDORA', 'Financiación'),
                ('2050', 'Obligaciones laborales', 'PASIVO', 'ACREEDORA', 'Operación'),
                ('3010', 'Capital social', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3020', 'Superávit de capital', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3030', 'Utilidades acumuladas', 'CAPITAL', 'ACREEDORA', 'Financiación'),
                ('3040', 'Resultado del período', 'CAPITAL', 'ACREEDORA', 'Operación'),
                ('4010', 'Ventas de productos', 'INGRESO', 'ACREEDORA', 'Operación'),
                ('4020', 'Devoluciones en ventas', 'INGRESO', 'DEUDORA', 'Operación'),
                ('4030', 'Descuentos concedidos', 'INGRESO', 'DEUDORA', 'Operación'),
                ('4040', 'Otros ingresos operativos', 'INGRESO', 'ACREEDORA', 'Operación'),
                ('5011', 'Costo - Materia prima directa', 'COSTO', 'DEUDORA', 'Operación'),
                ('5012', 'Costo - Mano de obra directa', 'COSTO', 'DEUDORA', 'Operación'),
                ('5013', 'Costos indirectos de fabricación', 'COSTO', 'DEUDORA', 'Operación'),
                ('5020', 'Gastos de operación', 'GASTO', 'DEUDORA', 'Operación'),
                ('5030', 'Gastos de venta', 'GASTO', 'DEUDORA', 'Operación'),
                ('5040', 'Gastos administrativos', 'GASTO', 'DEUDORA', 'Operación'),
                ('5050', 'Gastos financieros', 'GASTO', 'DEUDORA', 'Operación/Financiación'),
                ('5060', 'Depreciación y amortización', 'GASTO', 'DEUDORA', 'Operación (Ajuste no efectivo)'),
            ]
        }
    }
    
    if tipo_plantilla not in PLANTILLAS:
        return Response({'error': 'Tipo de plantilla no válido'}, status=400)
    
    # Obtener empresa actual
    empresa = None
    if hasattr(request, 'empresa') and request.empresa:
        empresa = request.empresa
    else:
        empresa_id = request.session.get('empresa_id')
        if empresa_id:
            from apps.empresas.models import Empresa
            try:
                empresa = Empresa.objects.get(id=empresa_id)
            except Empresa.DoesNotExist:
                pass
        
        if not empresa:
            from apps.empresas.models import UsuarioEmpresa
            acceso = UsuarioEmpresa.objects.filter(
                usuario=request.user,
                activo=True
            ).first()
            if acceso:
                empresa = acceso.empresa
    
    if not empresa:
        return Response({'error': 'Sin empresa seleccionada'}, status=400)
    
    plantilla = PLANTILLAS[tipo_plantilla]
    
    try:
        with transaction.atomic():
            # Eliminar todas las cuentas existentes
            CuentaContable.objects.filter(empresa=empresa).delete()
            
            # Crear las cuentas de la plantilla
            cuentas_creadas = []
            for codigo, nombre, tipo, naturaleza, categoria in plantilla['cuentas']:
                # Determinar nivel por el código
                nivel = len(codigo.replace('.', ''))
                
                cuenta = CuentaContable.objects.create(
                    empresa=empresa,
                    codigo=codigo,
                    nombre=nombre,
                    tipo=tipo,
                    naturaleza=naturaleza,
                    nivel=nivel,
                    afectable=True,  # Todas las cuentas afectables por defecto
                    creado_por=request.user,
                    modificado_por=request.user
                )
                cuentas_creadas.append(cuenta)
            
            return Response({
                'mensaje': f'Plantilla {plantilla["nombre"]} aplicada exitosamente',
                'cuentas_creadas': len(cuentas_creadas),
                'tipo': tipo_plantilla
            })
            
    except Exception as e:
        return Response({'error': f'Error al aplicar plantilla: {str(e)}'}, status=500)